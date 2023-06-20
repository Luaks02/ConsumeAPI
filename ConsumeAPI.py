from openpyxl import workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import os

def buscar_sistema():
    #Abrindo a planilha de controle
    wb = load_workbook(filename="Smart_Haus.xlsm", read_only=False,keep_vba=True)
    ws = wb["Dados_Tecnico-Projeto"]
    ws2 = wb["Dados_Proposta"]

    site = ws["B1"].value

    #Colhendo informações com Selenium

    #driver_path = r"C:\Users\Avell\Documents\ConsumeAPI\chrome-win64\chrome-win64\chrome.exe"
    #brave_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"

    options = Options()
    #options.binary_location = brave_path
    options.add_argument('--headless')
    options.add_argument("--window-size=1920,1200")
    #service = Service(driver_path)


    driver = webdriver.Chrome(options=options)
    driver.get(site)
    driver.implicitly_wait(20)
    Ok_cookie = driver.find_element(By.XPATH, '//button[text()="Estou ciente"]')
    Ok_cookie.click()
    driver.implicitly_wait(20)
    rowshoplist = driver.find_element(By.XPATH, "//div[@class='row shoplist']/ul/*")
    kit_info = driver.find_element(By.XPATH, "//div[@class='row shoplist']/ul/div").text
    rowshoplist.click()
    driver.implicitly_wait(5)
    description = driver.find_elements(By.XPATH, "//div[@id='description']")[0].text

    #Filtrando informações colhidas
    listed_description = []

    listed_description = description.splitlines()

    for index,desc in enumerate(listed_description):
        if desc[0:4] == "Área":
            if desc[28] == "²":
                area = desc[25:26]
                print("OK")
            else:
                area = desc[25:28]
        if desc[0:10] == "Peso sobre":
            peso = desc[23:25]
        if desc[0:36] == "O gerador de energia fotovoltaico de":
            inicio = index+1
        if desc[0:4] == "ESTE":
            fim = index-2
            break
        if desc[0:14] == "Regulamentação":
            fim = index-2

    listed_kit_info = []
    listed_kit_info = kit_info.splitlines()

    valor = listed_kit_info[3].split() 
    valor = str(valor)[4:-2]

    kwp = listed_kit_info[2].split()
    kwp = kwp[1][:-3]


    #Enviando informações para a planilha de controle

    numero = 3

    ws["B2"].value = "Equipamentos"
    ws["C2"].value = "Qnt"

    if ws2["F1"].value is None:
        ws2["F1"].value = valor

    ws2["C2"].value = kwp

    for linha in range(inicio,fim):
        material = listed_description[linha].split()
        ws["B" + str(numero)].value = " ".join(material[1:])
        ws["C" + str(numero)].value = int(material[0])

        if material[3] == "GROWATT":
            ws2["H29"].value = material[3]
            ws2["H31"].value = int(material[0])
            ws2["I29"].value = material[6]
            try:
                ws2["I31"].value = int(material[11][0])
            except:
                ws2["I31"].value = int(material[10][0])
            if material[7] == "E":
                ws2["J29"].value = material[8]
            else:
                ws2["J29"].value = material[7]
            numero += 1
            continue

        if material[3] == "FRONIUS":
            ws2["H29"].value = material[3]
            ws2["H31"].value = int(material[0])
            ws2["I29"].value = " ".join(material[4:6])
            ws2["I31"].value = int(material[9])
            ws2["J29"].value = material[6]
            numero += 1
            continue

        if material[3] == "DEYE":
            ws2["H29"].value = material[3]
            ws2["H31"].value = int(material[0])
            ws2["I29"].value = (material[4])
            ws2["I31"].value = int(material[8][0])
            ws2["J29"].value = material[5] 
            numero += 1
            continue

        if material[3] == "JINKO":
            ws2["H25"].value = material[3]
            ws2["H27"].value = int(material[0])
            ws2["I25"].value = material[4]
            ws2["J25"].value = material[7]
            numero += 1
            continue

        if material[3] == "JA":
            ws2["H25"].value = material[3]
            ws2["H27"].value = int(material[0])
            ws2["I25"].value = material[4]
            ws2["J25"].value = material[5]
            numero += 1
            continue

        numero += 1

    ws["B" + str(numero)].value = "Instalação"
    ws["C" + str(numero)].value = 1
    ws2["I27"].value = int(area)
    ws2["J27"].value = int(area)*int(peso)

    #Fechando sistemas

    wb.save("Smart_Haus.xlsm")
    driver.quit()

    #Abrindo planilha

    os.startfile("Smart_Haus.xlsm")

    print("Feche e salve a planilha primeiro. Agora responda R para repetir o script() ou responda F para fechar. =)")

    user_input = input()

    if user_input == "r":
        buscar_sistema()
    else:
        exit()

buscar_sistema()
 